# -*- coding: utf-8 -*-
"""crm_export.py — экспорт данных CRM в форматированный Excel (openpyxl)."""

from __future__ import annotations

import io
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BRAND_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="111111")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(*(Side(style="thin", color="CCCCCC"),) * 4)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 45)


def export_sales_xlsx(sales_rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"

    headers = ["Дата продажи", "Клиент", "Телефон", "Email", "Тип товара", "Модель",
               "Город доставки", "Скидка, %", "Цена, ₽"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    total = 0.0
    for row in sales_rows:
        ws.append([row.get("sale_date", ""), row.get("name_short", ""), row.get("phone", ""),
                   row.get("email", ""), row.get("product_type", ""), row.get("product_model", ""),
                   row.get("delivery_city", ""), row.get("discount_pct", 0), row.get("price", 0)])
        total += float(row.get("price", 0) or 0)
        for cell in ws[ws.max_row]:
            cell.border = THIN_BORDER

    ws.append(["", "", "", "", "", "", "", "Итого:", total])
    last_row = ws[ws.max_row]
    last_row[7].font = TOTAL_FONT
    last_row[8].font = TOTAL_FONT
    last_row[8].fill = BRAND_FILL

    _autosize(ws)
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_quotes_xlsx(quotes_rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    headers = ["Номер КП", "Дата", "Клиент", "Телефон", "Тип товара", "Модель", "Монтаж",
               "Город доставки", "Сумма без скидки, ₽", "Скидка, %", "Итог, ₽", "Статус", "Менеджер"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    from crm_db import STATUS_LABELS
    for row in quotes_rows:
        ws.append([row.get("kp_number", ""), row.get("created_at", ""), row.get("name_short", ""),
                   row.get("phone", ""), row.get("product_type", ""), row.get("product_model", ""),
                   "Да" if row.get("include_montage") else "Нет", row.get("delivery_city", ""),
                   row.get("base_total", 0), row.get("discount_pct", 0), row.get("final_total", 0),
                   STATUS_LABELS.get(row.get("status"), row.get("status")), row.get("owner_name", "")])
        for cell in ws[ws.max_row]:
            cell.border = THIN_BORDER

    _autosize(ws)
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
