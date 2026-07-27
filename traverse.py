"""Модуль вакуумных траверс VacuTec — прайс, парсер, комплектация, фото."""
from __future__ import annotations
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional
import pandas as pd

APP_DIR = Path(__file__).resolve().parent
TRAVERSE_MEDIA_DIR = APP_DIR / "media" / "traverse"
TRAVERSE_PRICE_FILE = "Vakuumnye-traversy-7.xlsx"

# Ставка НДС — совпадает с общей в приложении
VAT_RATE = 0.22

# Карта модель → файл фото
IMAGE_MAP = {
    "VacuTec 4P-350-L2.2-D200": "4P-350-4.jpg",
    "VacuTec 6P-500-L2.2-D200": "6P-500-4.jpg",
    "VacuTec 8P-760-L2.2-D200": "8P-760-2.jpg",
    "VacuTec 6P-1200-L2.2-D300": "6P-1200-5.jpg",
    "VacuTec 4PR-200-L2.2-D200": "VacuTec-4PR-200-L2.2-D200-1-2.jpg",
    "VacuTec 8PR-400-L2.2-D200": "VacuTec-8PR-400-L2.2-D200-2-4.jpg",
    "VacuTec 8PR-800-L2.2-D300": "VacuTec-8PR-800-L2.2-D300-1-5.jpg",
    "VacuTec 4GT-140-L2.2-D300": "VacuTec-4GT-140-L2.2-D300.jpg",
    "VacuTec 8GT-280-L2.2-D300": "VacuTec-8GT-280-L2.2-D300-3.jpg",
    "VacuTec 10P-900-L5.5-D200": "10P-900-9.jpg",
    "VacuTec 14P-1200-L5.5-D200": "14P-1200-7.jpg",
    "VacuTec 10P-2000-L5.5-D300": "10P-2000-10.jpg",
    "VacuTec 14P-2800-L5.5-D300": "14P-2800-8.jpg",
    "VacuTec 10PR-500-L5.5-D200": "VacuTec-10PR-500-L5.5-D200-2-1-6.jpg",
}

# Заглушка (когда для модели нет отдельного фото) — берём ближайшую по типу
FALLBACK_MAP = {
    "VacuTec 4P-800-L2.2-D300": "6P-1200-5.jpg",       # 4P с D300 — берём 6P-1200
    "VacuTec 8P-1600-L2.2-D300": "8P-760-2.jpg",
    "VacuTec 8PR-800-L5.5-D300": "VacuTec-8PR-800-L2.2-D300-1-5.jpg",
    "VacuTec 10PR-1000-L5.5-D300": "VacuTec-10PR-500-L5.5-D200-2-1-6.jpg",
    "VacuTec 12PR-1000-L5.5-D300": "VacuTec-10PR-500-L5.5-D200-2-1-6.jpg",
}

CRANE_PLUS_TRAVERSE_IMG = "Konsolnyi-kran-vakuumnaia-traversa-5.jpg"
PNEUMO_IMG = "Vakuumnaia-traversa-na-pnevmotsilindre-6.jpg"

# Опциональные позиции (цены с НДС)
OPTIONS = {
    "battery": {
        "key": "battery",
        "label": "Сменный аккумулятор 24V, 20,8Ah с блоком питания",
        "code": "VT-ACC-BAT",
        "price_novat": 39401.64,
    },
    "supports": {
        "key": "supports",
        "label": "Опоры для хранения траверсы типа Р",
        "code": "VT-ACC-SUP",
        "price_novat": 7500.0,
    },
    "cable": {
        "key": "cable",
        "label": "Спиральный кабель управления",
        "code": "VT-ACC-CBL",
        "price_novat": 15700.82,
    },
    "handle": {
        "key": "handle",
        "label": "Наклонная ручка",
        "code": "VT-ACC-HND",
        "price_novat": 17622.95,
    },
}


@dataclass
class TraverseSelection:
    """Выбор траверсы пользователем в UI."""
    base_code: str                    # напр. "VacuTec 6P-500-L2.2-D200"
    power_type: str = "220В"          # "220В" | "АКБ"
    options_battery: bool = False
    options_supports: bool = False
    options_cable: bool = False
    options_handle: bool = False
    include_montage: bool = False
    montage_price: float = 0.0
    montage_vat: bool = False
    prepay_pct: int = 100             # % предоплаты (по умолчанию 100)
    delivery_address: str = ""        # адрес/способ доставки
    delivery_price: float = 0.0       # стоимость доставки ₽ (если >0 — в спецификацию)

    @property
    def code(self) -> str:
        """Итоговый код с учётом суффикса A для АКБ-версии."""
        if self.power_type == "АКБ":
            return self.base_code + "A"
        return self.base_code


def load_traverse_price(uploaded=None) -> Optional[pd.DataFrame]:
    """Загружает прайс вакуумных траверс. Ищет в prices/, media/, корне."""
    import openpyxl
    from io import BytesIO

    if uploaded is not None:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
    else:
        for cand in [APP_DIR / "prices" / TRAVERSE_PRICE_FILE,
                     APP_DIR / TRAVERSE_PRICE_FILE]:
            if cand.exists():
                wb = openpyxl.load_workbook(cand, data_only=True)
                break
        else:
            return None

    ws = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):
        capacity = ws.cell(row=r, column=3).value
        code = ws.cell(row=r, column=4).value
        weight = ws.cell(row=r, column=8).value
        price_novat = ws.cell(row=r, column=9).value
        cutout = ws.cell(row=r, column=5).value
        suckers = ws.cell(row=r, column=6).value
        d_sucker = ws.cell(row=r, column=7).value
        if not code:
            continue
        code_s = str(code).strip()
        # Пропускаем разделитель "6м"
        if code_s.lower() in ("", "6м"):
            continue
        try:
            price_val = float(price_novat) if price_novat else 0.0
        except (TypeError, ValueError):
            price_val = 0.0
        rows.append({
            "code": code_s,
            "capacity": int(capacity) if isinstance(capacity, (int, float)) else None,
            "weight": str(weight) if weight else "",
            "price_novat": price_val,
            "price_vat": price_val * (1 + VAT_RATE),
            "cutout": str(cutout).replace("\n", " ") if cutout else "",
            "suckers": int(suckers) if isinstance(suckers, (int, float)) else None,
            "d_sucker": int(d_sucker) if isinstance(d_sucker, (int, float)) else None,
        })
    return pd.DataFrame(rows)


CODE_RE = re.compile(
    r"^VacuTec\s+(\d+)(P|PR|GT)-(\d+)-L(\d+\.\d+)-D(\d+)(A?)$"
)


def parse_code(code: str) -> Optional[dict]:
    """Разбирает код модели траверсы."""
    m = CODE_RE.match(code.strip())
    if not m:
        return None
    return {
        "suckers": int(m.group(1)),
        "type": m.group(2),
        "capacity": int(m.group(3)),
        "beam_length": float(m.group(4)),
        "d_sucker": int(m.group(5)),
        "is_battery": m.group(6) == "A",
        "base_code": code.replace("A", "") if code.endswith("A") else code,
    }


def strip_battery_suffix(code: str) -> str:
    """Убирает суффикс A (если есть)."""
    if code.endswith("A"):
        return code[:-1]
    return code


def get_traverse_image(base_code: str) -> Optional[Path]:
    """Возвращает путь к фото модели или подобранной заглушке."""
    base_code = strip_battery_suffix(base_code)
    fname = IMAGE_MAP.get(base_code) or FALLBACK_MAP.get(base_code)
    if not fname:
        return None
    p = TRAVERSE_MEDIA_DIR / fname
    return p if p.exists() else None


def get_cover_image(has_crane: bool, traverse_code: str) -> Optional[Path]:
    """Фото для титула КП: комбо или конкретная модель."""
    if has_crane and traverse_code:
        p = TRAVERSE_MEDIA_DIR / CRANE_PLUS_TRAVERSE_IMG
        return p if p.exists() else None
    return get_traverse_image(traverse_code)


def get_base_codes(df: pd.DataFrame) -> list[str]:
    """Список уникальных базовых кодов (без суффикса A)."""
    if df is None or df.empty:
        return []
    return sorted(set(strip_battery_suffix(c) for c in df["code"]))


def find_traverse_row(df: pd.DataFrame, code: str) -> Optional[pd.Series]:
    """Строка прайса по точному коду."""
    if df is None or df.empty:
        return None
    hit = df[df["code"] == code]
    if hit.empty:
        return None
    return hit.iloc[0]


def build_traverse_items(sel: TraverseSelection,
                         df: pd.DataFrame) -> list[dict]:
    """Собирает спецификацию для КП на траверсу.

    Возвращает список: [{code, name, unit, qty, price_vat, total_vat}]
    (цены — с НДС 22 %, как в остальных модулях приложения).
    """
    items: list[dict] = []
    row = find_traverse_row(df, sel.code)
    if row is not None:
        # Наименование: «Траверса вакуумная» + модель из прайса (суффикс A = версия на АКБ)
        items.append({
            "code": sel.code,
            "name": f"Траверса вакуумная {sel.code}",
            "unit": "шт",
            "qty": 1,
            "price_vat": float(row["price_vat"]),
            "total_vat": float(row["price_vat"]),
        })
    # Опции
    def add_opt(key: str, checked: bool):
        if not checked:
            return
        o = OPTIONS[key]
        p_vat = o["price_novat"] * (1 + VAT_RATE)
        items.append({
            "code": o["code"],
            "name": o["label"],
            "unit": "шт",
            "qty": 1,
            "price_vat": p_vat,
            "total_vat": p_vat,
        })

    add_opt("battery", sel.options_battery)
    add_opt("supports", sel.options_supports)
    add_opt("cable", sel.options_cable)
    add_opt("handle", sel.options_handle)
    return items


def build_traverse_quote_data(sel: TraverseSelection, df):
    """Собирает QuoteData-аналог для траверсы, чтобы использовать генерацию PDF/DOCX."""
    from dataclasses import dataclass, field

    @dataclass
    class TvQuote:
        series: str = "VACUTEC"
        capacity: int = 0
        boom: float = 0
        height_to_arm: float = 0
        hoist_brand: str = ""
        hoist_mode: str = ""
        hoist_height: int = 0
        include_electrification: bool = False
        include_montage: bool = False
        montage_price: float = 0.0
        montage_vat: bool = False
        hc_capacity = None
        column_diameter = None
        use_lllm: bool = False
        lllm_code = None
        include_flange: bool = False
        flange_code = None
        lines: list = field(default_factory=list)
        electrification_lines: list = field(default_factory=list)

        @property
        def total(self) -> float:
            return sum(l.total for l in self.lines)

    from dataclasses import dataclass as _dc

    @_dc
    class _SpecLine:
        code: str
        name: str
        unit: str
        qty: float
        price: float
        @property
        def total(self):
            return self.qty * self.price

    q = TvQuote()
    parsed = parse_code(sel.code) or {}
    q.capacity = parsed.get("capacity", 0)
    q.include_montage = sel.include_montage
    q.montage_price = sel.montage_price
    q.montage_vat = sel.montage_vat

    for item in build_traverse_items(sel, df):
        q.lines.append(_SpecLine(
            item["code"], item["name"], item["unit"],
            item["qty"], item["price_vat"]))
    return q


# Описания типов конструкции (профессиональный текст для КП)
TYPE_INTRO = {
    "P": "Вакуумная траверса VacuTec серии «P» — универсальное грузозахватное "
         "устройство для подъёма и перемещения листового металла со структурной "
         "поверхностью. Двухконтурные присоски уверенно удерживают лист толщиной "
         "от 0,5 мм и обеспечивают надёжный захват материала с покрытием эмульсией или следами коррозии.",
    "PR": "Вакуумная траверса VacuTec серии «PR» — усиленное грузозахватное "
          "устройство для подъёма и перемещения листового материала и длинномерных "
          "заготовок большого веса. Двухконтурные присоски обеспечивают надёжный "
          "захват листа толщиной от 0,5 мм, включая материал со следами эмульсии или коррозии.",
    "GT": "Вакуумная траверса VacuTec серии «GT» — грузозахватное устройство "
          "для подъёма и перемещения листов увеличенных габаритов. Двухконтурные "
          "присоски гарантируют надёжный захват материала толщиной от 0,5 мм.",
}

# Плоскость перемещения по серии
MOTION_PLANE = {
    "P":  "горизонтальная",
    "PR": "горизонтальная и вертикальная",
    "GT": "горизонтальная",
}

# Длина поперечной балки по диаметру присосок
CROSS_BEAM_MM = {200: 990, 300: 1290}


def get_max_sheet_length(sel: TraverseSelection) -> str:
    """Максимальная длина листа — по длине балки."""
    p = parse_code(sel.code) or {}
    bl = p.get("beam_length", 2.2)
    return f"{bl:g} м"


def get_traverse_description(sel: TraverseSelection, df) -> list[str]:
    """Подробное описание траверсы с подстановкой переменных из выбранной модели.

    Возвращает список: [вводная_фраза, пунктбулет_1, пунктбулет_2, ...].
    Вводная фраза — обычный абзац. Остальные — с булетами в PDF.
    """
    p = parse_code(sel.code) or {}
    suckers = p.get("suckers", "—")
    ttype = p.get("type", "P")
    beam_length = p.get("beam_length", 2.2)
    d_sucker = p.get("d_sucker", 200)
    capacity = p.get("capacity", 0)
    cross_beam = CROSS_BEAM_MM.get(d_sucker, 990)
    beam_mm = int(beam_length * 1000)
    max_sheet = f"{beam_length:g} м"

    intro = TYPE_INTRO.get(ttype, TYPE_INTRO["P"])

    bullets = [
        f"Главная балка из стали, длиной {beam_mm} мм.",
        "<b>Несущая балка выполняет функцию ресивера</b> для повышения безопасности работы.",
        f"Поперечные балки из стали, длиной {cross_beam} мм, перемещаются вдоль главной балки бесступенчато.",
        "Отключение присосок по отдельности — отсечные краны на каждой присоске (опция). "
        "Захват груза с минимальными размерами обеспечивается передвижением чаш вдоль поперечной балки.",
        "Присоски двухконтурные, немецкого производства.",
        "Вакуумный манометр с красно-зелёной шкалой в поле зрения оператора.",
        "Интеллектуальная электронная система мониторинга контролирует состояние системы и активирует "
        "акустическую и оптическую сигнализацию «ТРЕВОГА» при отклонении параметров от нормы.",
        "Лакокрасочное покрытие RAL 2008 (жёлтый глубокий).",
        "Поставляется в собранном виде, готовой к эксплуатации.",
    ]
    return [intro] + bullets


def traverse_characteristics(sel: TraverseSelection,
                             df: pd.DataFrame) -> list[tuple[str, str]]:
    """Характеристики для таблицы в КП."""
    row = find_traverse_row(df, sel.code)
    parsed = parse_code(sel.code) or {}
    ttype = parsed.get("type", "P")
    rows = [
        ("Модель", sel.code),
        ("Тип конструкции",
         {"P": "стандартная прямая", "PR": "усиленная/раздвижная",
          "GT": "для листов большой ширины"}.get(ttype, "—")),
        ("Грузоподъёмность", f"{parsed.get('capacity', '—')} кг"),
        ("Количество присосок", str(parsed.get("suckers", "—"))),
        ("Диаметр присоски", f"{parsed.get('d_sucker', '—')} мм"),
        ("Длина балки", f"{parsed.get('beam_length', '—')} м"),
        ("Плоскость перемещения",
         MOTION_PLANE.get(ttype, "горизонтальная")),
        ("Тип питания",
         "Аккумуляторная батарея 24 В, 20,8 А·ч"
         if sel.power_type == "АКБ" else "сеть 220 В"),
        ("Вакуумный насос", "10 м³/ч"),
    ]
    if row is not None:
        if row.get("cutout"):
            rows.append(("Размер заготовки max/min, мм", row["cutout"]))
        if row.get("weight"):
            rows.append(("Вес устройства", str(row["weight"]) + " кг"))
    rows.append(("Гарантия", "12 месяцев"))
    return rows
